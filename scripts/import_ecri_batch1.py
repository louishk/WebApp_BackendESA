"""
Import the hand-built Batch1 ECRI xlsx into ecri_batches / ecri_batch_ledgers.

This is a one-off historical import: the xlsx was produced outside the tool
(manual review), and we want it to live in the tool the way a normal batch
would so that analytics, outcomes tracking, and the LSETUP test flow can use
it as their source batch.

Usage:
    cd /home/louis/PycharmProjects/WebApp_BackendESA
    # Dry validation (no writes):
    python scripts/import_ecri_batch1.py --check
    # Real import:
    python scripts/import_ecri_batch1.py --created-by louis

Idempotent: refuses to re-import the same source file (detected via
ecri_batches.group_config->>'imported_from_xlsx_source').
"""

import argparse
import sys
from datetime import date, datetime, timedelta
from decimal import Decimal
from pathlib import Path
from uuid import uuid4

REPO_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO_ROOT / 'backend' / 'python'))

import openpyxl
from sqlalchemy import create_engine, text
from sqlalchemy.orm import sessionmaker

from common.config_loader import get_database_url
from common.ecri_dates import compute_effective_date, next_lease_anniversary
from common.models import ECRIBatch, ECRIBatchLedger

XLSX_PATH = REPO_ROOT / 'output' / 'ecri_batch_2026-04-08 - Copy.xlsx'
SHEET_NAME = 'ecri_batch_2026-04-08 final'
BATCH_NAME = 'Batch1 2026-04-08 (imported)'
NOTICE_PERIOD_DAYS = 14
TARGET_INCREASE_PCT = Decimal('12.00')  # representative; per-row pct is authoritative

REQUIRED_COLS = [
    'site_id', 'ledger_id', 'tenant_id', 'tenant_name', 'unit_id', 's_unit',
    'actual_rent', 'ecri_new_rent', 'ecri_increase_applied_pct',
    'ecri_increase_needed', 'control_group', 'moved_in_date',
    'rent_last_changed', 'tenure_months', 'ecri_eligible', 'Excluded',
    'Exclusion Reason', 'currency',
]


def _num(v):
    if v is None or v == '':
        return None
    return Decimal(str(v))


def _as_date(v):
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    return None


def load_rows():
    if not XLSX_PATH.exists():
        raise SystemExit(f"xlsx not found: {XLSX_PATH}")

    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        raise SystemExit(f"sheet {SHEET_NAME!r} missing; have: {wb.sheetnames}")
    ws = wb[SHEET_NAME]

    rows_iter = ws.iter_rows(values_only=True)
    headers = list(next(rows_iter))
    missing = [c for c in REQUIRED_COLS if c not in headers]
    if missing:
        raise SystemExit(f"xlsx missing required columns: {missing}")

    idx = {h: i for i, h in enumerate(headers)}
    parsed = []
    for row in rows_iter:
        if row is None or all(v is None for v in row):
            continue
        parsed.append({k: row[idx[k]] for k in REQUIRED_COLS})
    return parsed


def build_ledger(batch_id, row, notice_date, effective_date, paid_thru=None, anniv=None, bucket=None, next_lad=None):
    old_rent = _num(row['actual_rent'])
    new_rent = _num(row['ecri_new_rent'])
    pct = _num(row['ecri_increase_applied_pct'])
    amt = _num(row['ecri_increase_needed'])

    excluded = bool(row.get('Excluded'))
    exclusion_reason = row.get('Exclusion Reason')

    if old_rent is None or new_rent is None:
        return None  # row unusable, skip

    if pct is None:
        pct = Decimal('0')
    if amt is None:
        amt = (new_rent - old_rent) if (old_rent is not None and new_rent is not None) else Decimal('0')

    api_status = 'pending'
    api_response = None
    if excluded:
        api_status = 'skipped'
        api_response = {'reason': 'Excluded in xlsx', 'detail': exclusion_reason}
    elif not row.get('ecri_eligible'):
        api_status = 'skipped'
        api_response = {'reason': 'Not ECRI-eligible per xlsx'}

    return ECRIBatchLedger(
        batch_id=batch_id,
        site_id=int(row['site_id']),
        ledger_id=int(row['ledger_id']),
        tenant_id=int(row['tenant_id']) if row['tenant_id'] is not None else None,
        unit_id=int(row['unit_id']) if row['unit_id'] is not None else None,
        unit_name=str(row['s_unit']) if row['s_unit'] is not None else None,
        tenant_name=str(row['tenant_name']) if row['tenant_name'] is not None else None,
        control_group=int(row['control_group']) if row['control_group'] is not None else 0,
        old_rent=old_rent,
        new_rent=new_rent,
        increase_pct=pct,
        increase_amt=amt,
        planned_new_rent=new_rent,
        planned_increase_pct=pct,
        planned_increase_amt=amt,
        currency=str(row['currency']).strip().upper() if row.get('currency') else 'SGD',
        notice_date=notice_date,
        effective_date=effective_date,
        paid_thru_date=paid_thru,
        next_lad=next_lad,
        bucket=bucket,
        moved_in_date=_as_date(row.get('moved_in_date')),
        last_increase_date=_as_date(row.get('rent_last_changed')),
        tenure_months=int(row['tenure_months']) if row.get('tenure_months') is not None else None,
        api_status=api_status,
        api_response=api_response,
    )


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--check', action='store_true', help='Validate xlsx, no DB writes')
    ap.add_argument('--created-by', default='import_script', help='ecri_batches.created_by value')
    args = ap.parse_args()

    print(f"Reading {XLSX_PATH} …")
    rows = load_rows()
    print(f"  parsed {len(rows)} data rows")

    site_ids = sorted({int(r['site_id']) for r in rows if r['site_id'] is not None})
    excluded_count = sum(1 for r in rows if r.get('Excluded'))
    ineligible_count = sum(1 for r in rows if not r.get('ecri_eligible'))
    print(f"  sites: {len(site_ids)} ({site_ids})")
    print(f"  excluded flagged: {excluded_count}")
    print(f"  ecri_eligible=False: {ineligible_count}")

    engine = create_engine(get_database_url('pbi'), pool_pre_ping=True)
    Session = sessionmaker(bind=engine)
    session = Session()

    # Pre-fetch dPaidThru and dAnniv from ccws_ledgers for billing-cycle dates
    site_ids = sorted({int(r['site_id']) for r in rows if r['site_id'] is not None})
    ledger_ids_list = [int(r['ledger_id']) for r in rows if r['ledger_id'] is not None]
    anniv_rows = session.execute(text(
        'SELECT "SiteID", "LedgerID", "dPaidThru", "dAnniv" FROM ccws_ledgers '
        'WHERE "SiteID" = ANY(:sids) AND "LedgerID" = ANY(:lids)'
    ), {'sids': site_ids, 'lids': ledger_ids_list}).fetchall()
    anniv_map = {}
    for ar in anniv_rows:
        pt = ar[2].date() if ar[2] is not None else None
        an = ar[3].date() if ar[3] is not None else None
        anniv_map[(int(ar[0]), int(ar[1]))] = (pt, an)

    today = date.today()

    if args.check:
        # Dry-build ledgers to surface row-level errors
        built = 0
        unusable = 0
        for r in rows:
            pt, an = anniv_map.get((int(r['site_id']), int(r['ledger_id'])), (None, None)) if r['site_id'] and r['ledger_id'] else (None, None)
            eff, ntc, bkt = compute_effective_date(an, pt, today, NOTICE_PERIOD_DAYS)
            led = build_ledger(uuid4(), r, ntc, eff, paid_thru=pt, anniv=an, bucket=bkt,
                               next_lad=next_lease_anniversary(an, today) if an else None)
            if led is None:
                unusable += 1
            else:
                built += 1
        print(f"  buildable ledgers: {built}")
        print(f"  unusable rows (missing rent): {unusable}")
        print("OK (--check). No DB changes.")
        session.close()
        return

    source_marker = str(XLSX_PATH)
    existing = session.query(ECRIBatch).filter(
        text("group_config->>'imported_from_xlsx_source' = :src")
    ).params(src=source_marker).first()
    if existing:
        session.close()
        raise SystemExit(
            f"Already imported: batch_id={existing.batch_id} created_at={existing.created_at}"
        )

    batch_id = uuid4()

    batch = ECRIBatch(
        batch_id=batch_id,
        name=BATCH_NAME,
        site_ids=site_ids,
        target_increase_pct=TARGET_INCREASE_PCT,
        control_group_enabled=True,
        group_config={
            'imported_from_xlsx': True,
            'imported_from_xlsx_source': source_marker,
            'imported_at': datetime.utcnow().isoformat(),
            'sheet': SHEET_NAME,
        },
        total_ledgers=0,  # updated below
        status='draft',
        created_by=args.created_by,
        min_tenure_months=12,
        notice_period_days=NOTICE_PERIOD_DAYS,
        discount_reference_pct=Decimal('40.00'),
        attribution_window_days=90,
        notes='Imported from hand-built xlsx review of the first ECRI batch.',
    )
    session.add(batch)
    session.flush()

    ledgers = []
    skipped_unusable = 0
    for r in rows:
        pt, an = anniv_map.get((int(r['site_id']), int(r['ledger_id'])), (None, None)) if r['site_id'] and r['ledger_id'] else (None, None)
        eff, ntc, bkt = compute_effective_date(an, pt, today, NOTICE_PERIOD_DAYS)
        nxt_lad = next_lease_anniversary(an, today) if an else None
        led = build_ledger(batch_id, r, ntc, eff, paid_thru=pt, anniv=an, bucket=bkt, next_lad=nxt_lad)
        if led is None:
            skipped_unusable += 1
            continue
        ledgers.append(led)

    session.bulk_save_objects(ledgers)
    batch.total_ledgers = len(ledgers)
    session.commit()

    print(f"Imported batch_id={batch_id}")
    print(f"  ledgers inserted: {len(ledgers)}")
    print(f"  unusable rows skipped (missing rent): {skipped_unusable}")
    print(f"  status: draft")


if __name__ == '__main__':
    main()
