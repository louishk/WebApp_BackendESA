#!/usr/bin/env python3
"""
Populate unit_range_mappings table on esa_pbi from Excel Master sheet.

Reads temp/UnitsTypeDescriptionSummary.xlsx:
- SiteConfig sheet: Facility -> SiteCode mapping
- Master sheet (rows 5-199): unit range definitions with climate/NOKE info

Handles suffix-bearing unit refs like 5000C, 5000D-5000V by splitting
into suffix-specific ranges so the view can match precisely.

Usage:
    python -m scripts.populate_unit_range_mappings
    python -m scripts.populate_unit_range_mappings --dry-run
"""

import argparse
import re
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent))

import openpyxl
from sqlalchemy import create_engine, text
from common.config_loader import get_database_url


# Facility -> SiteCode mapping (from SiteConfig sheet)
FACILITY_SITE_MAP = {
    'WD': 'L017',
    'TP': 'L022',
    'BK': 'L002',
    'CW': 'L029',
    'CW1': 'L028',
    'EL': 'L003',
    'HV': 'L025',
    'IMM': 'L001',
    'WC': 'L004',
    'AMK': 'L018',
    'MMR': 'L005',
    'TS': 'L030',
    'KW': 'L008',
    'ESKY': 'L006',
    'ESKB': 'L011',
    'ESKA': 'L019',
    'ESKG': 'L013',
    'YDP': 'L021',
    'YS': 'L023',
    'BP': 'L024',
    'CSL': 'L007',
    'SGT': 'L009',
    'S51A': 'L010',
    'KD': 'L026',
}


def parse_unit_ref(raw: str) -> tuple:
    """
    Parse a unit reference like 'A30001', '5000C', '2998 &2998A', '1183M'
    into (prefix, number, suffix).

    Returns (prefix_str, number_int, suffix_str) or (None, None, None).
    """
    if raw is None:
        return None, None, None

    raw = str(raw).strip()
    if not raw:
        return None, None, None

    # Handle compound refs like '2998 &2998A' - take first part
    raw = raw.split('&')[0].split(',')[0].strip()

    # Extract leading letters as prefix
    m = re.match(r'^([A-Za-z]*)', raw)
    prefix = m.group(1).upper() if m and m.group(1) else ''

    # Extract first contiguous digits
    m_num = re.search(r'(\d+)', raw)
    if not m_num:
        return None, None, None

    number = int(m_num.group(1))

    # Extract trailing letters after digits as suffix
    remaining = raw[m_num.end():]
    suffix_match = re.match(r'^([A-Za-z]+)', remaining)
    suffix = suffix_match.group(1).upper() if suffix_match else ''

    return prefix, number, suffix


def load_excel(excel_path: str) -> list:
    """
    Load unit range mappings from Excel file.
    Returns list of dicts ready for DB insert.

    When a range like '5000C TO 5178' spans different numbers AND the start
    has a suffix, it is split into:
      - 5000-5000 suffix_start=C suffix_end=C  (the specific suffixed unit)
      - 5001-5178 suffix_start=NULL suffix_end=NULL  (the rest of the range)

    When start and end share the same number with suffixes (e.g. 5000D-5000V),
    the suffix range is stored directly.
    """
    wb = openpyxl.load_workbook(excel_path, data_only=True)

    if 'Master' not in wb.sheetnames:
        print(f"ERROR: 'Master' sheet not found. Available: {wb.sheetnames}")
        sys.exit(1)

    ws = wb['Master']
    records = []
    skipped = 0

    for row_idx in range(5, 200):  # rows 5-199 (1-indexed in openpyxl)
        facility = ws.cell(row=row_idx, column=1).value  # Col A
        range_start_raw = ws.cell(row=row_idx, column=4).value  # Col D
        range_end_raw = ws.cell(row=row_idx, column=6).value    # Col F
        storage_type = ws.cell(row=row_idx, column=7).value     # Col G
        climate_type = ws.cell(row=row_idx, column=8).value     # Col H
        dehumidifier = ws.cell(row=row_idx, column=9).value     # Col I
        noke_status = ws.cell(row=row_idx, column=10).value     # Col J

        # Skip empty rows
        if not facility or not climate_type:
            skipped += 1
            continue

        facility = str(facility).strip()
        site_code = FACILITY_SITE_MAP.get(facility)
        if not site_code:
            print(f"  WARN row {row_idx}: Unknown facility '{facility}', skipping")
            skipped += 1
            continue

        # Parse range start
        prefix_start, num_start, suffix_start = parse_unit_ref(range_start_raw)
        if num_start is None:
            print(f"  WARN row {row_idx}: Cannot parse range_start '{range_start_raw}', skipping")
            skipped += 1
            continue

        # Parse range end
        _, num_end, suffix_end = parse_unit_ref(range_end_raw)
        if num_end is None:
            num_end = num_start
            suffix_end = suffix_start

        # Normalize values
        climate_type = str(climate_type).strip()
        storage_type = str(storage_type).strip() if storage_type else None
        has_dehumidifier = str(dehumidifier).strip().upper() == 'YES' if dehumidifier else False
        noke = str(noke_status).strip() if noke_status else 'NO'

        base = {
            'site_code': site_code,
            'facility': facility,
            'unit_prefix': prefix_start,
            'storage_type': storage_type,
            'climate_type': climate_type,
            'has_dehumidifier': has_dehumidifier,
            'noke_status': noke,
        }

        if num_start == num_end and (suffix_start or suffix_end):
            # Same number with suffixes: e.g. 5000D-5000V -> suffix range
            records.append({
                **base,
                'range_start': num_start,
                'range_end': num_end,
                'suffix_start': suffix_start or None,
                'suffix_end': suffix_end or None,
            })
        elif num_start != num_end and suffix_start:
            # Different numbers but start has suffix: e.g. 5000C-5178
            # Split into suffix-specific entry + remaining range
            records.append({
                **base,
                'range_start': num_start,
                'range_end': num_start,
                'suffix_start': suffix_start,
                'suffix_end': suffix_start,
            })
            records.append({
                **base,
                'range_start': num_start + 1,
                'range_end': num_end,
                'suffix_start': None,
                'suffix_end': None,
            })
        else:
            # Normal range, no meaningful suffix
            records.append({
                **base,
                'range_start': num_start,
                'range_end': num_end,
                'suffix_start': None,
                'suffix_end': None,
            })

    print(f"  Parsed {len(records)} records, skipped {skipped} rows")
    return records


def populate(records: list, dry_run: bool = False):
    """Insert records into unit_range_mappings on esa_pbi."""
    if dry_run:
        print(f"\n  [DRY RUN] Would upsert {len(records)} records")
        for r in records[:15]:
            suffix_info = ''
            if r.get('suffix_start'):
                suffix_info = f" suffix={r['suffix_start']}-{r.get('suffix_end', r['suffix_start'])}"
            print(f"    {r['facility']} ({r['site_code']}): "
                  f"{r['unit_prefix']}{r['range_start']}-{r['range_end']}{suffix_info} "
                  f"-> {r['climate_type']}, NOKE={r['noke_status']}")
        if len(records) > 15:
            print(f"    ... and {len(records) - 15} more")
        return

    db_url = get_database_url('pbi')
    engine = create_engine(db_url)

    with engine.connect() as conn:
        # Verify table exists
        exists = conn.execute(text("""
            SELECT EXISTS (
                SELECT FROM information_schema.tables
                WHERE table_name = 'unit_range_mappings'
            )
        """)).scalar()

        if not exists:
            print("ERROR: unit_range_mappings table does not exist.")
            print("Run migration first: psql -f backend/python/migrations/010_unit_range_mappings.sql")
            sys.exit(1)

        # Clear existing data and re-insert
        conn.execute(text("DELETE FROM unit_range_mappings"))

        inserted = 0
        for r in records:
            conn.execute(text("""
                INSERT INTO unit_range_mappings
                    (site_code, facility, unit_prefix, range_start, range_end,
                     suffix_start, suffix_end,
                     storage_type, climate_type, has_dehumidifier, noke_status)
                VALUES
                    (:site_code, :facility, :unit_prefix, :range_start, :range_end,
                     :suffix_start, :suffix_end,
                     :storage_type, :climate_type, :has_dehumidifier, :noke_status)
            """), r)
            inserted += 1

        conn.commit()
        print(f"\n  Inserted {inserted} records into unit_range_mappings")

        # Summary by site
        result = conn.execute(text("""
            SELECT facility, site_code, count(*) as ranges,
                   count(*) FILTER (WHERE suffix_start IS NOT NULL) as suffix_ranges
            FROM unit_range_mappings
            GROUP BY facility, site_code
            ORDER BY facility
        """))
        print("\n  Summary:")
        for row in result:
            suffix_note = f" ({row.suffix_ranges} suffix-specific)" if row.suffix_ranges else ""
            print(f"    {row.facility:6s} ({row.site_code}): {row.ranges} ranges{suffix_note}")


def main():
    parser = argparse.ArgumentParser(
        description="Populate unit_range_mappings from Excel Master sheet"
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Preview without writing to database"
    )
    parser.add_argument(
        "--excel",
        default=None,
        help="Path to Excel file (default: temp/UnitsTypeDescriptionSummary.xlsx)"
    )
    args = parser.parse_args()

    # Find Excel file
    if args.excel:
        excel_path = Path(args.excel)
    else:
        base = Path(__file__).parent.parent.parent.parent  # project root
        excel_path = base / 'temp' / 'UnitsTypeDescriptionSummary.xlsx'

    if not excel_path.exists():
        print(f"ERROR: Excel file not found: {excel_path}")
        sys.exit(1)

    print("=" * 60)
    print("Populate unit_range_mappings from Excel")
    print("=" * 60)
    print(f"  Excel: {excel_path}")
    if args.dry_run:
        print("  Mode:  DRY RUN")

    print("\nParsing Excel...")
    records = load_excel(str(excel_path))

    if not records:
        print("No records parsed. Exiting.")
        sys.exit(1)

    print("\nPopulating database...")
    populate(records, dry_run=args.dry_run)

    print("\n" + "=" * 60)
    print("Done.")
    print("=" * 60)


if __name__ == "__main__":
    main()
